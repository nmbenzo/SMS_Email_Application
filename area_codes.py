import openpyxl
import re

wb = openpyxl.load_workbook(
    '/Users/nbenzschawel/Downloads/area_codes_by_state.xlsx')
ws = wb.active

area_codes = []

def build_area_codes_list():

    for rowNum in range(2, ws.max_row):
        code = ws.cell(row=rowNum, column=1).value
        area_codes.append(code)
    return area_codes

build_area_codes_list()

intl_num = []
loc_num = []


def filter_intl_phone(area_codes):
    user_input = '+205-915488'
    cleaned_phone = re.sub(r'[^\w]', '', user_input)
    for x in area_codes:
        if x == str(cleaned_phone[0:3]):
            loc_num.append(cleaned_phone)
        else:
            intl_num.append(cleaned_phone)




filter_intl_phone(area_codes)
print(intl_num)
print(loc_num)
