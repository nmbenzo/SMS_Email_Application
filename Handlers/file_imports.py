import openpyxl


location_a = '/Users/nbenzschawel/Library/Mobile Documents/' \
             'com~apple~CloudDocs/Comms_Application'
location_b = '/Users/nbenzschawel/Downloads'


""" SEVIS Active Student Data """

active_stud_issm_data = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/All_SEVIS-Active_Student_Tracking.xlsx'

active_student_req_reg = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Raw_Files/SEVIS - Active Students Requiring Registration.xlsx'

ACTIVE_students_FINAL = '/Users/nbenzschawel/Downloads/SEVIS_Reg/2019/Spring/Final Workbooks/ACTIVE_Final.xlsx'


ACTIVE_wb = openpyxl.Workbook()

try:
	ACTIVE_students = openpyxl.load_workbook(ACTIVE_students_FINAL)
	ACTIVE_students_sheet = ACTIVE_students.worksheets[0]
except FileNotFoundError:
	ACTIVE_wb.save(ACTIVE_students_FINAL)
	ACTIVE_students = openpyxl.load_workbook(ACTIVE_students_FINAL)
	ACTIVE_students_sheet = ACTIVE_students.worksheets[0]

wb2_active = openpyxl.load_workbook(active_student_req_reg)
sheet = wb2_active.worksheets[0]
ws2 = wb2_active.active

wb1 = openpyxl.load_workbook(active_stud_issm_data)
ws = wb1.active

